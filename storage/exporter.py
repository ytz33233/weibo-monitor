import json
import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import SentimentRecord
from storage.database import Database
import config

logger = logging.getLogger(__name__)

ZEBRA_FILL_1 = PatternFill(fill_type="solid", fgColor="FFFFFF")
ZEBRA_FILL_2 = PatternFill(fill_type="solid", fgColor="F7F9FC")
KPI_FILL = PatternFill(fill_type="solid", fgColor="EAF2FF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
KPI_FONT = Font(name="Arial", bold=True, size=11, color="1F2937")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9DEE7"),
    right=Side(style="thin", color="D9DEE7"),
    top=Side(style="thin", color="D9DEE7"),
    bottom=Side(style="thin", color="D9DEE7"),
)


def _apply_zebra(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        fill = ZEBRA_FILL_1 if (r - min_row) % 2 == 0 else ZEBRA_FILL_2
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).fill = fill
            ws.cell(r, c).border = THIN_BORDER


def _write_header(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def export_excel(db: Database, output_path: str = None) -> str:
    if not output_path:
        os.makedirs(config.EXPORT_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(config.EXPORT_DIR, f"舆情监控_{timestamp}.xlsx")

    wb = Workbook()

    # Sheet 1: 汇总概览
    ws_summary = wb.active
    ws_summary.title = "汇总概览"
    summary = db.get_summary()
    keyword_stats = db.get_keyword_stats()

    ws_summary.cell(1, 1, "工行活动舆情监控 — 汇总概览").font = Font(name="Arial", bold=True, size=14)
    ws_summary.merge_cells("A1:F1")

    kpi_labels = ["总帖子数", "今日新增", "总互动量", "微博帖子", "小红书帖子", "最后采集时间"]
    kpi_values = [
        summary["total"], summary["today_count"], summary["total_interactions"],
        summary["weibo_count"], summary["xhs_count"],
        summary["last_collected"] or "暂无数据",
    ]
    for i, (label, value) in enumerate(zip(kpi_labels, kpi_values)):
        row = 3 + i
        ws_summary.cell(row, 1, label).font = KPI_FONT
        ws_summary.cell(row, 1).fill = KPI_FILL
        ws_summary.cell(row, 1).border = THIN_BORDER
        ws_summary.cell(row, 2, value).font = BODY_FONT
        ws_summary.cell(row, 2).border = THIN_BORDER

    start_row = 10
    ws_summary.cell(start_row, 1, "关键词统计").font = Font(name="Arial", bold=True, size=12)
    stat_headers = ["关键词", "平台", "帖子数", "总点赞", "总评论", "总转发/分享", "总收藏"]
    _write_header(ws_summary, start_row + 1, stat_headers, [15, 10, 10, 12, 12, 15, 12])
    for i, stat in enumerate(keyword_stats):
        row = start_row + 2 + i
        vals = [stat["keyword"], stat["platform"], stat["count"],
                stat["total_likes"], stat["total_comments"],
                stat["total_shares"], stat["total_collects"]]
        for j, v in enumerate(vals, 1):
            ws_summary.cell(row, j, v).font = BODY_FONT
    _apply_zebra(ws_summary, start_row + 2, start_row + 1 + len(keyword_stats), 1, 7)
    ws_summary.column_dimensions["A"].width = 15

    # Sheet 2: 微博明细
    ws_weibo = wb.create_sheet("微博明细")
    weibo_headers = ["关键词", "标题", "作者", "点赞", "评论", "转发", "发布时间", "采集时间", "内容"]
    _write_header(ws_weibo, 1, weibo_headers, [12, 35, 12, 8, 8, 8, 18, 18, 50])
    weibo_data = db.query(platform="weibo")
    for i, row_data in enumerate(weibo_data):
        r = 2 + i
        vals = [row_data["keyword"], row_data["title"], row_data["author"],
                row_data["likes"], row_data["comments"], row_data["shares"],
                row_data["created_at"], row_data["collected_at"], row_data["content"]]
        for j, v in enumerate(vals, 1):
            ws_weibo.cell(r, j, v).font = BODY_FONT
            ws_weibo.cell(r, j).alignment = Alignment(wrap_text=True, vertical="top")
    _apply_zebra(ws_weibo, 2, 1 + len(weibo_data), 1, 9)

    # Sheet 3: 小红书明细
    ws_xhs = wb.create_sheet("小红书明细")
    xhs_headers = ["关键词", "标题", "作者", "点赞", "评论", "分享", "收藏", "发布时间", "采集时间", "内容"]
    _write_header(ws_xhs, 1, xhs_headers, [12, 35, 12, 8, 8, 8, 8, 18, 18, 50])
    xhs_data = db.query(platform="xhs")
    for i, row_data in enumerate(xhs_data):
        r = 2 + i
        vals = [row_data["keyword"], row_data["title"], row_data["author"],
                row_data["likes"], row_data["comments"], row_data["shares"],
                row_data["collects"], row_data["created_at"], row_data["collected_at"],
                row_data["content"]]
        for j, v in enumerate(vals, 1):
            ws_xhs.cell(r, j, v).font = BODY_FONT
            ws_xhs.cell(r, j).alignment = Alignment(wrap_text=True, vertical="top")
    _apply_zebra(ws_xhs, 2, 1 + len(xhs_data), 1, 10)

    # Sheet 4: 趋势数据
    ws_trend = wb.create_sheet("趋势数据")
    trend_headers = ["日期", "平台", "帖子数", "互动量"]
    _write_header(ws_trend, 1, trend_headers, [15, 10, 10, 15])
    trend_data = db.get_daily_trend()
    for i, row_data in enumerate(trend_data):
        r = 2 + i
        vals = [row_data["date"], row_data["platform"], row_data["count"], row_data["interactions"]]
        for j, v in enumerate(vals, 1):
            ws_trend.cell(r, j, v).font = BODY_FONT
    _apply_zebra(ws_trend, 2, 1 + len(trend_data), 1, 4)

    wb.save(output_path)
    logger.info(f"Excel 导出成功: {output_path}")
    return output_path


def export_json(db: Database, output_path: str = None) -> str:
    if not output_path:
        os.makedirs(config.EXPORT_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(config.EXPORT_DIR, f"舆情监控_{timestamp}.json")

    all_data = db.query()
    export = {
        "export_time": datetime.now().isoformat(),
        "summary": db.get_summary(),
        "keyword_stats": db.get_keyword_stats(),
        "records": all_data,
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(export, f, ensure_ascii=False, indent=2, default=str)
    logger.info(f"JSON 导出成功: {output_path}")
    return output_path
